import os
import re
from typing import List, Dict, Set
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from models import Box, Area

def parse_floor_count(floor_str: str) -> int:
    """
    Parse floor string to get the number of layers.
    Examples:
    "1" -> 1
    "4-5" -> 2 (5-4+1)
    "6-34" -> 29 (34-6+1)
    "1,2,5-17" -> 1 + 1 + 13 = 15
    """
    if not floor_str:
        return 1
    total = 0
    parts = [p.strip() for p in str(floor_str).split(",") if p.strip()]
    if not parts:
        return 1
    for p in parts:
        # Range "X-Y"
        m = re.match(r"^(\d+)\s*-\s*(\d+)$", p)
        if m:
            start = int(m.group(1))
            end = int(m.group(2))
            total += abs(end - start) + 1
            continue
        # Single number
        m2 = re.match(r"^\d+$", p)
        if m2:
            total += 1
            continue
        # Fallback: count as 1
        total += 1
    return total

def export_to_excel(boxes: List[Box], output_dir: str, filename: str = "面积计算表.xlsx", checker: str = "", reviewer: str = ""):
    """
    Export boxes to Excel file in output_dir.
    """
    if not boxes:
        print("No boxes to export.")
        return

    filepath = os.path.join(output_dir, filename)
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Group boxes by Building
    buildings: Dict[str, List[Box]] = {}
    for box in boxes:
        b_name = box.Building
        if b_name not in buildings:
            buildings[b_name] = []
        buildings[b_name].append(box)
    
    # Sort buildings to keep order (optional, but good)
    sorted_buildings = sorted(buildings.keys())
    
    for b_name in sorted_buildings:
        b_boxes = buildings[b_name]
        
        create_summary_sheet(wb, b_name, b_boxes, checker, reviewer)
        create_detail_sheet(wb, b_name, b_boxes, checker, reviewer)
        
    wb.save(filepath)
    print(f"Excel file saved to: {filepath}")

def create_detail_sheet(wb, building_name, boxes: List[Box], checker: str = "", reviewer: str = ""):
    """
    Create "单体X" sheet with summary format.
    Format example:
     单体建筑面积统计表 
     栋号 	 名  称 	 建筑面积 （平方米） 	 备注 
     3# 	 住宅 	 6055.19 	 含装配式建筑面积103.20平方米 
     	     商业
     ...
     合           计 	 	 6055.19 
     注： ...
    """
    sheet_name = f"单体{building_name}"
    ws = wb.create_sheet(title=sheet_name)
    
    # 1. Aggregate data across all boxes (floors) for this building
    # Need to sum up areas by name/type
    # Also need to track total building area, basal area, etc.
    
    jirong_map: Dict[str, float] = {}
    bujirong_map: Dict[str, float] = {}
    total_area = 0.0
    basal_area = 0.0 # Only from first floor usually, or max? Assuming sum of '基底' areas if present, or maybe just look for '基底' in areas.
    # Actually, basal area usually refers to the footprint area. In the box data, do we have "基底"?
    # In previous logic, we looked for areas with "基底" in name/attr.
    
    # Collect all area names for consistent ordering
    all_jirong_names = set()
    all_bujirong_names = set()
    
    for box in boxes:
        count = parse_floor_count(box.Floor)
        
        # Check for basal area (usually on floor 1 or implicitly defined)
        # If any area is marked as "基底", we use it.
        # But wait, previous logic summed basal areas.
        for area in box.areas:
            val = area.value * count
            
            if "基底" in area.name or "基底" in area.elseAttribute:
                # If basal is explicitly an area, add it to basal_area, but NOT to total_area usually?
                # "建筑基底面积" is usually separate.
                basal_area += val
                continue # Don't add to total area if it's just footprint
            
            if "计容" in area.elseAttribute:
                if "不计容" not in area.elseAttribute:
                    jirong_map[area.name] = jirong_map.get(area.name, 0.0) + val
                    total_area += val
                    all_jirong_names.add(area.name)
                else:
                    bujirong_map[area.name] = bujirong_map.get(area.name, 0.0) + val
                    total_area += val
                    all_bujirong_names.add(area.name)
            elif "不计容" in area.elseAttribute:
                bujirong_map[area.name] = bujirong_map.get(area.name, 0.0) + val
                total_area += val
                all_bujirong_names.add(area.name)
    
    # Define Rows
    # Jirong items first, then Bujirong items
    # Standard order if possible
    jirong_special = ["住宅", "商业"]
    jirong_others = sorted([n for n in all_jirong_names if n not in jirong_special])
    jirong_ordered = [n for n in jirong_special if n in all_jirong_names] + jirong_others
    
    bujirong_special = ["避难层", "架空层", "物业管理用房", "消防控制室", "社区服务用房", "养老服务用房"]
    bujirong_others = sorted([n for n in all_bujirong_names if n not in bujirong_special])
    # Merge sorted lists based on priority
    bujirong_ordered = [n for n in bujirong_special if n in all_bujirong_names] + bujirong_others
    
    # Also need to include items from user example even if 0? 
    # User example has empty cells for some items. 
    # "商业", "养老...", "物业...", "消防...", "社区...", "避难层", "架空层"
    # I should list them if they are in the standard list, even if value is 0?
    # User example: "3# 住宅 6055.19 ... 商业 (empty) ..."
    # So yes, list standard items.
    
    standard_items = ["住宅", "商业", "养老服务用房", "物业管理用房", "消防控制室", "社区服务用房", "避难层", "架空层"]
    # We should merge detected items with standard items to ensure all are listed
    # Use a list of dicts for rows
    
    rows = []
    # Add detected items that are NOT in standard list
    # Actually, let's just use the order: Residential, Commercial, then others from detected, then standard bujirong, then detected bujirong
    
    # Strategy: 
    # 1. Residential
    # 2. Commercial
    # 3. Other Jirong detected
    # 4. Standard Bujirong/Facilities (养老, 物业, 消防, 社区, 避难, 架空)
    # 5. Other Bujirong detected
    
    # Prepare data source
    # We need to look up value from jirong_map OR bujirong_map
    def get_val(name):
        return jirong_map.get(name, 0.0) + bujirong_map.get(name, 0.0)

    processed_names = set()
    
    # 1. Residential
    if "住宅" in all_jirong_names or "住宅" in all_bujirong_names:
        rows.append({"name": "住宅", "val": get_val("住宅")})
        processed_names.add("住宅")
    else:
        # If not present, maybe still add if it's a standard template? 
        # User example has "住宅" with value.
        # If 3# has only commercial, should we list Residential?
        # User example seems to list all standard categories.
        # Let's add it with empty value if not present?
        # The user input shows "商业" empty. So yes.
        rows.append({"name": "住宅", "val": 0.0}) # Will be empty string
        processed_names.add("住宅")

    # 2. Commercial
    rows.append({"name": "商业", "val": get_val("商业")})
    processed_names.add("商业")
    
    # 3. Other Jirong
    for name in jirong_ordered:
        if name not in processed_names:
            rows.append({"name": name, "val": get_val(name)})
            processed_names.add(name)
            
    # 4. Standard Facilities
    facilities = ["养老服务用房", "物业管理用房", "消防控制室", "社区服务用房", "避难层", "架空层"]
    for fac in facilities:
        rows.append({"name": fac, "val": get_val(fac)})
        processed_names.add(fac)
        
    # 5. Other Bujirong
    for name in bujirong_ordered:
        if name not in processed_names:
            rows.append({"name": name, "val": get_val(name)})
            processed_names.add(name)
            
    # Remove duplicates if any (though logic should prevent)
    
    # Start writing
    
    # Row 1: Title
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value="单体建筑面积统计表")
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.font = Font(size=16, bold=True)
    
    # Row 2: Empty
    
    # Row 3: Headers
    headers = ["栋号", "名  称", "建筑面积\n（平方米）", "备注"]
    ws.append(headers)
    # Style headers
    for i in range(1, 5):
        c = ws.cell(row=3, column=i)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
    
    # Data Rows
    start_row = 4
    current_row = start_row
    
    for idx, item in enumerate(rows):
        # Col 1: 栋号 (Only on first row)
        if idx == 0:
            c = ws.cell(row=current_row, column=1, value=f"{building_name}#")
            c.alignment = Alignment(horizontal='center', vertical='center')
        else:
            c = ws.cell(row=current_row, column=1, value="")
            
        # Col 2: Name
        c = ws.cell(row=current_row, column=2, value=item["name"])
        c.alignment = Alignment(horizontal='center', vertical='center')
        
        # Col 3: Value
        val = item["val"]
        if val != 0:
            c = ws.cell(row=current_row, column=3, value=val)
            c.number_format = '0.00'
        else:
            c = ws.cell(row=current_row, column=3, value="")
        c.alignment = Alignment(horizontal='center', vertical='center')
        
        # Col 4: Remark
        # Example: "含装配式建筑面积103.20平方米" for Residential?
        # Leave empty for now or customize?
        c = ws.cell(row=current_row, column=4, value="")
        c.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 1
        
    # Total Row
    ws.cell(row=current_row, column=1, value="合           计")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2) # Merge "合计" across A and B? Example shows "合           计" in A, B empty? 
    # Example: "合           计" 	 	 6055.19
    # Usually "合计" spans name column too or is just in first column.
    # Let's merge A and B.
    ws.merge_cells(f"A{current_row}:B{current_row}")
    c = ws.cell(row=current_row, column=1)
    c.alignment = Alignment(horizontal='center', vertical='center')
    
    c = ws.cell(row=current_row, column=3, value=total_area)
    c.number_format = '0.00'
    c.alignment = Alignment(horizontal='center', vertical='center')
    
    c = ws.cell(row=current_row, column=4, value="")
    
    current_row += 1
    
    # Borders for table
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for r in range(3, current_row):
        for c_idx in range(1, 5):
            ws.cell(row=r, column=c_idx).border = border
            
    # Notes Section
    # Note 1
    ws.cell(row=current_row, column=1, value="注：")
    ws.cell(row=current_row+1, column=1, value="1、本次建筑面积按建筑外墙边线进行计算")
    
    # Note 2
    # 2、本栋总建筑面积：   平方米，计容建筑面积：  平方米，建筑基底面积：  平方米
    # We need to calculate Total Jirong and Basal
    total_jirong = sum(jirong_map.values())
    note2 = f"2、本栋总建筑面积：{total_area:.2f}平方米，计容建筑面积：{total_jirong:.2f}平方米，建筑基底面积：{basal_area:.2f}平方米"
    ws.cell(row=current_row+2, column=1, value=note2)
    
    # Note 3
    # 3、本栋建筑标准层高：xxx米 (Hardcoded as per template? Or extracted?)
    # Since we don't have height info, use placeholder or fixed value from example.
    ws.cell(row=current_row+3, column=1, value="3、本栋建筑标准层高：   米")
    
    # Note 4
    # 4、本栋建筑高度：xxx米...
    note4 = "4、本栋建筑高度：    米（起算点：□ ±0  ■ 室外地面(散水)  最高点：■ 顶层层面   □ 屋面女儿墙顶  □ 其他）"
    ws.cell(row=current_row+4, column=1, value=note4)
    
    # Footer
    # 填充空格以保持格式，或者直接使用 f-string
    # 原始格式: "校核者：            检查者：                          日期：  年  月  日"
    # 如果没有提供名字，保留空格以便手写
    
    checker_str = checker if checker else "            "
    reviewer_str = reviewer if reviewer else "                          "
    
    # 尝试保持一定的间距，如果名字很长可能需要调整
    # 简单拼接
    footer = f"校核者：{checker_str}      检查者：{reviewer_str}      日期：  年  月  日"
    ws.cell(row=current_row+5, column=1, value=footer)
    
    # Merge notes cells across columns for better visibility?
    # Example implies they span.
    for r in range(current_row, current_row+6):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')

    # Column Widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30

def create_summary_sheet(wb, building_name, boxes: List[Box], checker: str = "", reviewer: str = ""):
    """
    Create "统X" sheet with summary table.
    """
    sheet_name = f"统{building_name}"
    ws = wb.create_sheet(title=sheet_name)
    
    # 1. Analyze columns
    # We need to find all unique Area names for Jirong, MianJirong and Bujirong
    jirong_names = set()
    mianjirong_names = set()
    bujirong_names = set()
    
    for box in boxes:
        for area in box.areas:
            ea = area.elseAttribute or ""
            if ("计容" in ea) and ("不计容" not in ea) and ("免计容" not in ea):
                jirong_names.add(area.name)
            elif "免计容" in ea:
                mianjirong_names.add(area.name)
            elif "不计容" in ea:
                bujirong_names.add(area.name)
    
    # Define Column Order
    # Standard lists as per user requirement
    std_jirong = ["住宅", "商业", "社区文化活动中心", "社区医疗卫生中心", "消防控制室", "垃圾收集屋"]
    # Default for "其他(住宅配套)免计容面积"
    std_mianjirong = ["养老服务用房", "物业用房", "配电房"]
    std_bujirong = ["架空层", "地下室出地面建筑"]
    
    # Identify extra columns present in data but not in standard list
    extra_jirong = sorted([n for n in jirong_names if n not in std_jirong])
    extra_mianjirong = sorted([n for n in mianjirong_names if n not in std_mianjirong])
    extra_bujirong = sorted([n for n in bujirong_names if n not in std_bujirong])
    
    # Final Column Lists
    # Logic: Dynamic items inserted at front, fixed items reduced from end to maintain total count (if possible)
    
    # Jirong
    jirong_cols = list(extra_jirong)
    num_fixed_j = len(std_jirong) - len(extra_jirong)
    if num_fixed_j > 0:
        jirong_cols.extend(std_jirong[:num_fixed_j])
    # If dynamic items exceed fixed slots, we just use dynamic items (and length grows)
    
    # MianJirong
    mianjirong_cols = list(extra_mianjirong)
    num_fixed_mj = len(std_mianjirong) - len(extra_mianjirong)
    if num_fixed_mj > 0:
        mianjirong_cols.extend(std_mianjirong[:num_fixed_mj])
    
    # Bujirong
    bujirong_cols = list(extra_bujirong)
    num_fixed_b = len(std_bujirong) - len(extra_bujirong)
    if num_fixed_b > 0:
        bujirong_cols.extend(std_bujirong[:num_fixed_b])
    
    # Calculate Total Columns
    # Base columns: Floor, Count, Basal, Total, Jirong_Subtotal
    # For each Jirong col: if in ["住宅", "商业"], takes 2 cols (Single, Multi). Else 1 col.
    
    # Mapping column indices
    # Col 1: 楼层
    # Col 2: 层数
    # Col 3: 建筑基底面积
    # Col 4: 总计
    # Col 5: 计容面积 (Start) -> Subtotal
    
    col_defs = []
    col_defs.append({"type": "prop", "key": "Floor", "header1": "楼层", "width": 10})
    col_defs.append({"type": "prop", "key": "Count", "header1": "层数", "width": 6})
    col_defs.append({"type": "prop", "key": "Basal", "header1": "建筑\n基底\n面积", "width": 12})
    col_defs.append({"type": "calc", "key": "Total", "header1": "总   计", "width": 12})
    
    # Jirong Section
    jirong_start_idx = len(col_defs)
    col_defs.append({"type": "jirong_subtotal", "header2": "小计", "width": 12})
    
    for name in jirong_cols:
        if name in ["住宅", "商业"]:
            col_defs.append({"type": "jirong_item_single", "name": name, "header2": name, "header3": "单层", "width": 10})
            col_defs.append({"type": "jirong_item_multi", "name": name, "header2": name, "header3": "多层小计", "width": 12})
        else:
            col_defs.append({"type": "jirong_item_simple", "name": name, "header2": name, "width": 12})
            
    # MianJirong Section
    mianjirong_start_idx = len(col_defs)
    col_defs.append({"type": "mianjirong_subtotal", "header2": "小计", "width": 12})
    for name in mianjirong_cols:
        col_defs.append({"type": "mianjirong_item_simple", "name": name, "header2": name, "width": 12})
    
    bujirong_start_idx = len(col_defs)
    col_defs.append({"type": "bujirong_subtotal", "header2": "小计", "width": 12})
    
    for name in bujirong_cols:
        if name == "架空层":
            col_defs.append({"type": "bujirong_item_single", "name": name, "header2": name, "header3": "单层", "width": 10})
            col_defs.append({"type": "bujirong_item_multi", "name": name, "header2": name, "header3": "多层小计", "width": 12})
        else:
            col_defs.append({"type": "bujirong_item_simple", "name": name, "header2": name, "width": 12})
        
    col_defs.append({"type": "prop", "key": "Remark", "header1": "备注", "width": 10})
    
    # Write Headers
    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_defs))
    c1 = ws.cell(row=1, column=1, value="单栋面积统计表")
    c1.alignment = Alignment(horizontal='center', vertical='center')
    c1.font = Font(size=16, bold=True)
    
    # Row 2: Info
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(col_defs))
    c2 = ws.cell(row=2, column=1, value=f"栋号： {building_name}                                                 单位：米、平方米")
    c2.alignment = Alignment(horizontal='left', vertical='center')
    
    # Row 3, 4, 5: Table Headers
    # Base headers (Floor, Count, etc) span 3 rows (3-5)
    # Jirong Header spans from jirong_start_idx+1 to bujirong_start_idx
    # Bujirong Header spans from bujirong_start_idx+1 to len(col_defs)-1
    
    # Write "计容面积" and "不计容面积" in Row 3
    # Indices are 0-based in list, 1-based in Excel
    
    # Jirong Span
    j_start = jirong_start_idx + 1
    j_end = mianjirong_start_idx
    if j_end >= j_start:
        ws.merge_cells(start_row=3, start_column=j_start, end_row=3, end_column=j_end)
        c = ws.cell(row=3, column=j_start, value="计容面积")
        c.alignment = Alignment(horizontal='center', vertical='center')
        
    # MianJirong Span
    mj_start = mianjirong_start_idx + 1
    mj_end = bujirong_start_idx
    if mj_end >= mj_start:
        ws.merge_cells(start_row=3, start_column=mj_start, end_row=3, end_column=mj_end)
        c = ws.cell(row=3, column=mj_start, value="其他(住宅配套)免计容面积")
        c.alignment = Alignment(horizontal='center', vertical='center')
        
    # Bujirong Span
    b_start = bujirong_start_idx + 1
    b_end = len(col_defs) - 1 # Excluding Remark
    if b_end >= b_start:
        ws.merge_cells(start_row=3, start_column=b_start, end_row=3, end_column=b_end)
        c = ws.cell(row=3, column=b_start, value="不计容面积")
        c.alignment = Alignment(horizontal='center', vertical='center')

    # Iterate cols to write specific headers
    for i, col in enumerate(col_defs):
        c_idx = i + 1
        if col.get("header1"):
            # Spans Row 3-5
            ws.merge_cells(start_row=3, start_column=c_idx, end_row=5, end_column=c_idx)
            c = ws.cell(row=3, column=c_idx, value=col["header1"])
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif col.get("header2"):
            # Row 4
            # Check if merged with next (for Residential/Commercial split)
            if col["type"] in ("jirong_item_single", "bujirong_item_single"):
                # Merge with next in Row 4
                ws.merge_cells(start_row=4, start_column=c_idx, end_row=4, end_column=c_idx+1)
                c = ws.cell(row=4, column=c_idx, value=col["header2"])
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif col["type"] in ("jirong_item_multi", "bujirong_item_multi"):
                # Already handled by previous
                pass
            else:
                # Simple items or subtotals: Merge Row 4-5?
                # Example shows "小计" in Row 4, and Row 5 is empty/merged?
                # For "养老服务用房", Row 4 is name, Row 5 is empty?
                # Usually better to merge 4-5 if no sub-header.
                if "header3" not in col:
                    ws.merge_cells(start_row=4, start_column=c_idx, end_row=5, end_column=c_idx)
                    c = ws.cell(row=4, column=c_idx, value=col["header2"])
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    # Has header3
                    c = ws.cell(row=4, column=c_idx, value=col["header2"]) # Should be handled by merge logic above actually
                    pass
        
        if col.get("header3"):
            c = ws.cell(row=5, column=c_idx, value=col["header3"])
            c.alignment = Alignment(horizontal='center', vertical='center')

    # Data Rows
    current_row = 6
    
    # Totals accumulators
    col_totals = {i: 0.0 for i in range(len(col_defs))}
    
    for box in boxes:
        # Prepare row data
        count = parse_floor_count(box.Floor)
        
        # Collect areas
        areas_map = {} # name -> value
        # Distinguish Jirong/Bujirong
        jirong_areas = {}
        mianjirong_areas = {}
        bujirong_areas = {}
        
        total_jirong = 0.0
        total_mianjirong = 0.0
        total_bujirong = 0.0
        
        for area in box.areas:
            ea = area.elseAttribute or ""
            if ("计容" in ea) and ("不计容" not in ea) and ("免计容" not in ea):
                jirong_areas[area.name] = area.value
                total_jirong += area.value
            elif "免计容" in ea:
                mianjirong_areas[area.name] = area.value
                total_mianjirong += area.value
            elif "不计容" in ea:
                bujirong_areas[area.name] = area.value
                total_bujirong += area.value
        
        row_vals = []
        for i, col in enumerate(col_defs):
            val = ""
            ctype = col["type"]
            
            if ctype == "prop":
                if col["key"] == "Floor":
                    val = box.Floor
                elif col["key"] == "Count":
                    val = count
                elif col["key"] == "Basal":
                    # Look for "基底" area
                    # Try to find in all areas
                    val = 0.0
                    for a in box.areas:
                        if "基底" in a.name or "基底" in a.elseAttribute:
                            val += a.value
                    if val == 0: val = ""
                elif col["key"] == "Remark":
                    val = ""
            
            elif ctype == "calc":
                if col["key"] == "Total":
                    # Sum of Multi-layer totals for this row
                    val = round((total_jirong + total_mianjirong + total_bujirong) * count, 2)
            
            elif ctype == "jirong_subtotal":
                # Subtotal for this floor (Multi-layer)
                val = round(total_jirong * count, 2)
            
            elif ctype == "jirong_item_single":
                name = col["name"]
                v = jirong_areas.get(name, 0.0)
                val = v if v != 0 else ""
                
            elif ctype == "jirong_item_multi":
                name = col["name"]
                v = jirong_areas.get(name, 0.0)
                val = round(v * count, 2) if v != 0 else ""
                
            elif ctype == "jirong_item_simple":
                name = col["name"]
                v = jirong_areas.get(name, 0.0)
                val = round(v * count, 2) if v != 0 else "" # Simple items usually show total for the floor range
                
            elif ctype == "mianjirong_subtotal":
                val = round(total_mianjirong * count, 2)
                
            elif ctype == "mianjirong_item_simple":
                name = col["name"]
                v = mianjirong_areas.get(name, 0.0)
                val = round(v * count, 2) if v != 0 else ""
                
            elif ctype == "bujirong_subtotal":
                val = round(total_bujirong * count, 2)
                
            elif ctype == "bujirong_item_single":
                name = col["name"]
                v = bujirong_areas.get(name, 0.0)
                val = v if v != 0 else ""
                
            elif ctype == "bujirong_item_multi":
                name = col["name"]
                v = bujirong_areas.get(name, 0.0)
                val = round(v * count, 2) if v != 0 else ""
                
            elif ctype == "bujirong_item_simple":
                name = col["name"]
                v = bujirong_areas.get(name, 0.0)
                val = round(v * count, 2) if v != 0 else ""
            
            row_vals.append(val)
            
            # Accumulate totals
            if isinstance(val, (int, float)):
                col_totals[i] += val
        
        # Write Row
        for i, v in enumerate(row_vals):
            c = ws.cell(row=current_row, column=i+1, value=v)
            c.alignment = Alignment(horizontal='center', vertical='center')
            if isinstance(v, (int, float)):
                c.number_format = '0.00'
        
        current_row += 1
        
    # Total Row
    ws.cell(row=current_row, column=1, value="合计")
    for i in range(2, len(col_defs)): # Skip Floor and Count? Count total usually not needed or sum
        # Only sum value columns
        ctype = col_defs[i].get("type", "")
        if ctype not in ["prop"] or col_defs[i].get("key") == "Basal":
             val = col_totals[i]
             if val != 0:
                 c = ws.cell(row=current_row, column=i+1, value=val)
                 c.number_format = '0.00'
                 c.alignment = Alignment(horizontal='center', vertical='center')

    # Footer
    current_row += 1
    
    checker_str = checker if checker else "                    "
    reviewer_str = reviewer if reviewer else "                                                                                                       "
    
    footer_text = f"校核者：{checker_str}      检查者：{reviewer_str}      日期：  年  月  日"
    
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(col_defs))
    c = ws.cell(row=current_row, column=1, value=footer_text)
    
    # Borders
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # Apply borders to table part (Row 3 to Total Row)
    for r in range(3, current_row): # Exclude footer
        for c_idx in range(1, len(col_defs) + 1):
            ws.cell(row=r, column=c_idx).border = border
